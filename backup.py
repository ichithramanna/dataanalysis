import xml.etree.ElementTree as ET
import pandas as pd
import xlwt
from xlwt import Workbook
import os
import openpyxl
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from tkinter import filedialog
filename = askopenfilename(title="Select Input File")
#tree = ET.parse('C:/Users/ichithr/Desktop/Samples/audi.xml')
tree = ET.parse(filename)
root = tree.getroot()

#OUTPUT FILE NAME
var = filename.split('/')
x=var[-1]
outsrs=x.replace('.xml','_OUT_SRS')
outitp=x.replace('.xml','_OUT_ITP')
outhead=x.replace('.xml','_OUT_Header')
outc=x.replace('.xml','_OUT_C')
outmain=x.replace('.xml','_OUT_MAin')
caplmin=x.replace('.xml','_OUT_caplmin')
caplmid=x.replace('.xml','_OUT_caplmid')
caplmax=x.replace('.xml','_OUT_caplmax')
rx1=x.replace('.xml','_OUT_Rx1')
rx2=x.replace('.xml','_OUT_Rx2')
rx3=x.replace('.xml','_OUT_Rx3')
rx4=x.replace('.xml','_OUT_Rx4')


#OUTPUT FILE LOCATION
savelocation=filedialog.askdirectory(title="Select Current Release Folder")
OUTSRS= savelocation + '/' + outsrs + '.xls'
OUTITP=savelocation + '/' + outitp + '.xls'
OUTHEAD=savelocation + '/' + outhead + '.h'
OUTC=savelocation + '/' + outc + '.c'
OUTMAIN=savelocation + '/' + outmain + '.c'
CAPLMIN=savelocation + '/' + caplmin + '.c'
CAPLMID=savelocation + '/' + caplmid + '.c'
CAPLMAX=savelocation + '/' + caplmax + '.c'
RX1=savelocation + '/' + rx1 + '.c'
RX2=savelocation + '/' + rx2 + '.c'
RX3=savelocation + '/' + rx3 + '.c'
RX4=savelocation + '/' + rx4 + '.c'
caten=[]
dlc=[]


#CAN-FRAME IDENTIFIER
CanIdentifier = {}
names= []
canid =[]
identifier=[]
for name in root.findall(".//ELEMENTS/CAN-FRAME/SHORT-NAME"):
    names.append(name.text)
for child in root.findall(".//FRAME-TRIGGERINGS/CAN-FRAME-TRIGGERING"):
    canid.append(child.attrib['UUID'])
for ide in root.findall(".//FRAME-TRIGGERINGS/CAN-FRAME-TRIGGERING/IDENTIFIER"):
    identifier.append(ide.text)

for _ in range(len(identifier)):
    identifier[_]=hex(int(identifier[_]))



for _ in range(len(names)):
    CanIdentifier[names[_]] = identifier[_]
dc= pd.DataFrame.from_dict(CanIdentifier, orient='index',columns=['identifier'])
#print (dc)
caten=[]
dlc=[]
for child in root.findall(".//ELEMENTS/CAN-FRAME"):
    for pict in child.findall(".//SHORT-NAME"):
        for ttpt in names:
            if ttpt == pict.text:
                caten.append(pict.text)
                for tcip in child.findall(".//FRAME-LENGTH"):
                    dlc.append(tcip.text)
DLC=dict(zip(caten,dlc))

resul = {}
counter = 0
c = 0
array = {}
data =[]
new=[]
tp=[]
pbo=[]
pos=[]
sig =[]
tig={}
pig={}
noo =[]
val=[]
cbs=[]
oo=[]
yes=[]
a= []
b=[]
f=[]
length=[]
jig={}
signalname=[]
dl=[]
ktttk=[]
#print('new')
for child in root.findall(".//I-SIGNAL-I-PDU"):
        for clat in child.findall(".//SHORT-NAME"):
            yes.append(clat.text)
            for ttpt in names:
                if ttpt == clat.text:
                    print('yes')
                    ktttk.append(clat.text)
                    for play in child.findall(".//LENGTH"):
                        dl.append(play.text)


for child in root.findall(".//I-SIGNAL-I-PDU"):
       # noo.append(child.text)
        noo.append(child.attrib['UUID'])
        #print(noo[counter])
        for name in child.findall(".//I-SIGNAL-TO-I-PDU-MAPPING/SHORT-NAME"):
            sig.append(name.text)


#INIT-VALUE
for child in root.findall(".//INIT-VALUE"):
#    oo.append(child.attrib['UUID'])
    for valu in child.findall(".//VALUE"):
        val.append(valu.text)
    for al in child.findall(".//SHORT-LABEL"):
        v=al.text
        m=v.replace("Init_",'')
        f.append(m)
value=dict(zip(f,val))
kat=[]
ale=[]

#INVALID-VALUE
for child in root.findall(".//INVALID-VALUE"):
#    oo.append(child.attrib['UUID'])
    for alll in child.findall(".//VALUE"):
        ale.append(alll.text)
    for al in child.findall(".//SHORT-LABEL"):
        l=al.text
        n=l.replace("Invalid_",'')
        kat.append(n)

invalue=dict(zip(kat,ale))

#LENGTH
for child in root.findall(".//I-SIGNAL"):
    for come in child.findall(".//SHORT-NAME"):
        oo.append(come.text)
    for com in child.findall(".//LENGTH"):
        cbs.append(com.text)

leng=dict(zip(oo,cbs))
e=0
d=0

#START-POSITION,TRANSFER-PROPERTY,PACKING-BYTE-ORDER
for child in root.findall(".//I-SIGNAL-TO-I-PDU-MAPPING"):
    c=1
    d=1
    e=1
    for neighbor in child.iter('I-SIGNAL-REF'):
        n=neighbor.text
        j=n.replace("/ISignal/",'')
        signalname.append(j)
        c=0
    if c==1:
        signalname.append('NO SIGNAL')
    for sta in child.findall(".//START-POSITION") :
        pos.append(sta.text)
    for start in child.findall(".//TRANSFER-PROPERTY"):
        tp.append(start.text)
        d=0
    if d==1:
        tp.append('N/A')
    for star in child.findall(".//PACKING-BYTE-ORDER"):
        p=star.text
        j=p.replace("MOST-SIGNIFICANT-BYTE-LAST",'INTEL')
        pbo.append(j)
        e=0
    if e==1:
        pbo.append('N/A')

#TIME-OUT
x='NO SIGNAL'
port=[]
ciggi=[]
for child in root.findall(".//I-SIGNAL-TRIGGERING"):
    for kp in child.iter('I-SIGNAL-PORT-REF'):
        c=1
        n=kp.text
        j=n.replace("/ECU/TSG_FS/CN_HCP4_CANFD01/",'')
        j=j.replace("/ECU/TSG_HFS/CN_HCP4_CANFD01/",'')
        j=j.replace("/ECU/TSG_BFS/CN_HCP4_CANFD01/",'')
        j=j.replace("/ECU/TSG_HBFS/CN_HCP4_CANFD01/",'')
        j=j.replace("/ECU/TSG_FS/CN_TSG_FS_LIN01/",'')
        j=j.replace("/ECU/TSG_HFS/CN_TSG_FS_LIN01/",'')
        j=j.replace("/ECU/TSG_BFS/CN_TSG_FS_LIN01/",'')
        u=j.replace("/ECU/TSG_HBFS/CN_TSG_FS_LIN01/",'')
        port.append(u)
        for jp in child.iter('I-SIGNAL-REF'):
            l=jp.text
            m=l.replace("/ISignal/",'')
            ciggi.append(m)
            c=0
        if c==1:
            ciggi.append(x)

porttosig= dict(zip(ciggi,port))
timeo=[]
porto=[]
for child in root.findall(".//I-SIGNAL-PORT"):
    for kp in child.iter('TIMEOUT'):
        timeo.append(kp.text)
        for jp in child.iter('SHORT-NAME'):
            porto.append(jp.text)




timee= dict(zip(porto,timeo))
timeou={}
for a,b in timee.items():
    for c,d in porttosig.items():
        if a==d:
            timeou[c]=b

for name in root.findall(".//I-SIGNAL/LENGTH"):
    length.append(name.text)

tig= dict(zip(a,zip(pos,length,pbo,tp,val,cbs)))
lp= pd.DataFrame.from_dict(tig,orient='index')
#print(lp)
#lp.to_excel("C:/Users/ichithr/Desktop/out.xlsx")

tfer= pd.DataFrame.from_dict(pig, orient='index',columns=['POS'])


#print(pos)
ara=[]
para=[]
inv=[]
tio=[]

for i in signalname:
    for key,element in value.items():
        if key == i:
            count=1
            var=value[key]
    if count == 1:
        para.append(var)
        count=0
    else :
        para.append('N/A')

initvalue=dict(zip(signalname,para))

for i in signalname:
    for key,element in leng.items():
        if key == i:
            count=1
            var=leng[key]
    if count == 1:
        ara.append(var)
        count=0
    else :
        ara.append('N/A')

lengsignal=dict(zip(signalname,ara))

for i in signalname:
    for key,element in invalue.items():
        if key == i:
            count=1
            var=invalue[key]
    if count == 1:
        inv.append(var)
        count=0
    else :
        inv.append('N/A')

invalidvalue=dict(zip(signalname,inv))

count=0
for i in signalname:
    for key,element in timeou.items():
        if key == i:
            count=1
            var=timeou[key]
    if count == 1:
        tio.append(var)
        count=0
    else :
        tio.append('N/A')

timeoutee=dict(zip(signalname,tio))



#CAN-FRAME DIRECTION
ta=[]
pa=[]
for child in root.findall(".//CAN-FRAME-TRIGGERING"):
    for neighbor in child.iter('FRAME-PORT-REF'):
        n=neighbor.text
        j=n.replace("/ECU/TSG_FS/CN_HCP4_CANFD01/",'')
        ta.append(j)
    for neighbor in child.iter('FRAME-REF'):
        n=neighbor.text
        j=n.replace("/Frame/",'')
        pa.append(j)
direc= dict(zip(pa,ta))
sa=[]
for key in direc:
    y=direc[key]
    if 'Tx' in y:
        sa.append(key)

txrx={}
for a in pa:
#    for b in sa:
        if a in sa:
            txrx[a]='Tx'
        else:
            txrx[a]='Rx'


#TRANSMISSION-MODE-TRUE-TIMING
na=[]
to=[]
tpp=[]
nop=[]
top=[]
rp=[]
lala=[]
for child in root.findall(".//I-SIGNAL-I-PDU"):
    top.append('1')
    j=1
    k=1
    l=1
    s=1
    for alll in child.findall(".//SHORT-NAME"):
        x=alll.text
        for i in pa:
            if i == x:

                lala.append(x)
                for ppl in child.findall(".//TRANSMISSION-MODE-TRUE-TIMING"):

                    for apl in ppl.findall(".//TIME-OFFSET/VALUE"):
                        to.append(apl.text)
                        k=0
                    if k==1:
                        to.append('N/A')
                    for nal in ppl.findall(".//TIME-PERIOD/VALUE"):
                        tpp.append(nal.text)
                        l=0
                    if l==1:
                        tpp.append('N/A')
                    for pal in ppl.findall(".//NUMBER-OF-REPETITIONS"):
                        nop.append(pal.text)
                        j=0
                    if j==1:
                        nop.append('N/A')
                    for sal in ppl.findall(".//REPETITION-PERIOD/VALUE"):
                        rp.append(sal.text)
                        s=0
                    if s==1:
                        rp.append('N/A')





timeoff= dict(zip(lala,to))
timeper= dict(zip(lala,tpp))
noofrep= dict(zip(lala,nop))
repper= dict(zip(lala,rp))
x='N/A'
for name in names:
    c=1
    for key in timeoff:
        if name==key:
            c=0
    if c==1:
        timeoff[name]=x
for name in names:
    c=1
    for key in timeper:
        if name==key:
            c=0
    if c==1:
        timeper[name]=x
for name in names:
    c=1
    for key in noofrep:
        if name==key:
            c=0
    if c==1:
        noofrep[name]=x
for name in names:
    c=1
    for key in repper:
        if name==key:
            c=0
    if c==1:
        repper[name]=x
#ONLY FOR Tx
for a,b in txrx.items():
    if b=='Rx':
        timeoff[a]=x
        timeper[a]=x
        noofrep[a]=x
        repper[a]=x

#CAN-FRAME SIGNAL GROUP
grp={}
for child in root.findall(".//I-SIGNAL-GROUP"):
    for sta in child.findall(".//SHORT-NAME") :
        k=sta.text
    for neighbor in child.iter('I-SIGNAL-REF'):
        n=neighbor.text
        j=n.replace("/ISignal/",'')
        grp[j]=k

ct=[]
for i in signalname:
    for key,element in grp.items():
        if key == i:
            count=1
            var=grp[key]
    if count == 1:
        ct.append(var)
        count=0
    else :
        ct.append('N/A')

group= dict(zip(signalname,ct))


print('Properties')
print(len(signalname))
print(len(pos))
print(len(tp))
print(len(pbo))
print(len(ara))
print(len(para))
print(len(inv))
print(len(tio))
final= dict(zip(signalname,zip(ct,pos,pbo,tp,para,ara,inv,tio)))
print(len(final))
cane={}

for a in range(len(names)):
    for b in range(len(signalname)):
        if names[a] in signalname[b]:
            cane.setdefault(names[a],[]).append(signalname[b])
x=['NO SIGNAL']
for name in names:
    c=1
    for key in cane:
        if name==key:
            c=0
    if c==1:
        cane[name]=x


fc=pd.concat({k: pd.Series(v) for k, v in cane.items()})
fc.to_excel("out.xlsx")
#print(group)

#WRITE TO EXCEL FOR SRS
l=1000

l=0
wb=Workbook()
sheet1=wb.add_sheet('SRS')
style = xlwt.easyxf('font: bold 1')
sheet1.write(1, 0, 'CAN-FRAME',style)
sheet1.write(1, 1, 'CAN-ID',style)
sheet1.write(1, 2, 'DLC',style)
sheet1.write(1, 3, 'CAN-DIRECTION',style)
sheet1.write(1, 4, 'TIME-OFFSET(secs)',style)
sheet1.write(1, 5, 'TIME-PERIOD(secs)',style)
sheet1.write(1, 6, 'NUMBER-OF-REP(secs)',style)
sheet1.write(1, 7, 'REPETITION-PER',style)
sheet1.write(1, 8, 'SIGNAL_NAME',style)
sheet1.write(1, 9, 'SIGNAL_GROUP',style)
sheet1.write(1, 10, 'START_POSITION',style)
sheet1.write(1, 11, 'ENDIAN',style)
sheet1.write(1, 12, 'TRANSFER-PROPERTY',style)
sheet1.write(1, 13, 'INIT-VALUE',style)
sheet1.write(1, 14, 'LENGTH(bits)',style)
sheet1.write(1, 15, 'INVALID-VALUE',style)
sheet1.write(1, 16, 'TIMEOUT',style)
r=3


for keys, values in cane.items():
    l=0
    sheet1.write(r,l,keys)
    l+=1
    for a,b in CanIdentifier.items():
        if keys==a:
            sheet1.write(r,l,b)
            l+=1
    for a,b in DLC.items():
        if keys==a:
            sheet1.write(r,l,b)
            l+=1
    for a,b in txrx.items():
        if keys==a:
            sheet1.write(r,l,b)
            l+=1
    for a,b in timeoff.items():
        if keys==a:
            sheet1.write(r,l,b)
            l+=1
    for a,b in timeper.items():
        if keys==a:
            sheet1.write(r,l,b)
            l+=1
    for a,b in noofrep.items():
        if keys==a:
            sheet1.write(r,l,b)
            l+=1
    for a,b in repper.items():
        if keys==a:
            sheet1.write(r,l,b)
            l+=1
    for a,b in final.items():
        l=8
        for signal in values:
            if signal==a:
                sheet1.write(r,l,signal)
                l+=1
                for para in b:
                    sheet1.write(r,l,para)
                    l+=1
                r+=1





#savedto=s=r"TSG_HFS_OUT.xls"
wb.save(OUTSRS)
print(' SRS Output saved to Excel')

##wb = openpyxl.load_workbook('TSG_BFS_OUT_SRS.')
##sheet1 = wb.active
##sheet1.merge_cells('A3:A8')

#WRITE TO EXCEL FOR ITP
l=1000

l=0
wb=Workbook()
sheet1=wb.add_sheet('temp')
style = xlwt.easyxf('font: bold 1')
sheet1.write(1, 0, 'requirement ID',style)
sheet1.write(1, 1, 'revision',style)
sheet1.write(1, 2, 'testcase no',style)
sheet1.write(1, 3, 'feature',style)
sheet1.write(1, 4, 'precondition',style)
sheet1.write(1, 5, 'testcase',style)
sheet1.write(1, 6, 'expected CanIdentifier',style)

r=5

e=0

naam=[]
naam2=[]
naam3=[]
naam4=[]
naam6=[]
naam7=[]
naam10=[]
naam11=[]
testid=[]

##To test the Transmission of CAN frame for CAN ID,with Signal Min/Mid/Max values and with Periodicity
for a,b in txrx.items():
        if b=='Tx':
            for keys, values in CanIdentifier.items():
                if keys==a:
                    for c,d in timeper.items():
                        if keys==c:
                            if d!='N/A':
                                naam.append(c)
                                l=2
                                e+=1
                                testid.append(e)
                                sheet1.write(r,l,e)
                                l+=1
                                n="To test the Transmission of CAN frame for CAN ID ",values," with Signal Min/Mid/Max values and with Periodicity",d,"sec."
                                sheet1.write(r,l,n)
                                l+=1
                                sheet1.write(r,l,'Com stack is initialized.')
                                l+=1
                                j="Step1: Invoke Rte_Write_<p>_<o>(Data) for each signal to update Max. \nStep2: Verify that return value of Rte_Write_<p>_<o>(Data)  for each signal is RTE_E_OK ( 0x00).\nStep3: Verify the Transmitted Data on CANalyzer. \n where <p> is the port name and <o> the VariableDataPrototype within the sender-receiver interface categorizing the port."

                                sheet1.write(r,l,j)
                                l+=1
                                m="1. The CanIdentifier Variable should be updated with value 0 if the return value of  Rte_Write_<p>_<o>(Data) for each signal is verified as RTE_E_OK ( 0x00).\n2.  The CANalyzer should observe the CAN ID ",values,".\n3. The Data on CANalyzer should be observed as MIN.\n4. The Data on CANalyzer should be observed for every",d,"sec period."
                                sheet1.write(r,l,m)
                                r+=1

##To test the Reception of CAN frame for CAN ID,with signal Min/Mid/Max values and with Periodicity
for a,b in txrx.items():
        if b=='Rx':
            for keys, values in CanIdentifier.items():
                if keys==a:
                        naam2.append(a)
                        l=2
                        e+=1
                        testid.append(e)
                        sheet1.write(r,l,e)
                        l+=1
                        n="To test the Reception of CAN frame for CAN ID ",values," with signal Min/Mid/Max values and with Periodicity",d,"sec."
                        sheet1.write(r,l,n)
                        l+=1
                        sheet1.write(r,l,'Com stack is initialized.')
                        l+=1
                        j="Step1: Invoke Rte_Write_<p>_<o>(Data) for each signal to update Max. \nStep2: Verify that return value of Rte_Write_<p>_<o>(Data)  for each signal is RTE_E_OK ( 0x00).\nStep3: Verify the Transmitted Data on CANalyzer. \n where <p> is the port name and <o> the VariableDataPrototype within the sender-receiver interface categorizing the port."
                        sheet1.write(r,l,j)
                        l+=1
                        m="1. Capture snapshot of CANalyzer window with Valid Data transmitted to ECU.\n2. Capture snapshot for valid indication.\n3. The CanIdentifier Variable should be updated with  equivalent data values .\n4. The Data on ECU should be observed for only one time.\n5. Notification should be invoked for the configured signals and its respective notification."
                        sheet1.write(r,l,m)
                        r+=1

##To test the Transmission of CAN frame for CAN ID,with signal value when user requested from CANalyzer/CANoe.
for a,b in txrx.items():
        if b=='Tx':
            for keys, values in CanIdentifier.items():
                if keys==a:
                    for c,d in timeper.items():
                        if keys==c:
                            if d=='N/A':
                                naam3.append(c)
                                l=2
                                e+=1
                                testid.append(e)
                                sheet1.write(r,l,e)
                                l+=1
                                n="To test the Transmission of CAN frame for CAN ID ",values,"  with signal value when user requested from CANalyzer/CANoe."
                                sheet1.write(r,l,n)
                                l+=1
                                sheet1.write(r,l,'Com stack is initialized.')
                                l+=1
                                j="Step1: Invoke Rte_Write_<p>_<o>(Data) for each signal to update Max. \nStep2: Verify that return value of Rte_Write_<p>_<o>(Data)  for each signal is RTE_E_OK ( 0x00).\nStep3: Verify the Transmitted Data on CANalyzer. \n where <p> is the port name and <o> the VariableDataPrototype within the sender-receiver interface categorizing the port."

                                sheet1.write(r,l,j)
                                l+=1
                                m="1. The CanIdentifier Variable should be updated with value 0 if the return value of  Rte_Write_<p>_<o>(Data) for each signal is verified as RTE_E_OK ( 0x00).\n2.  TheCANalyzer should observe the CAN ID ",values,".\n3. The Data on CANalyzer should be observed as MIN.\n4. The Data on CANalyzer should be observed for every",d,"sec period."
                                sheet1.write(r,l,m)
                                r+=1

##To test initial values for each Transmission signals
for a,b in txrx.items():
        if b=='Tx':
            for keys, values in CanIdentifier.items():
                if keys==a:
                        naam4.append(a)
                        l=2
                        e+=1
                        testid.append(e)
                        sheet1.write(r,l,e)
                        l+=1
                        n="To test initial values for each Transmission signals"
                        sheet1.write(r,l,n)
                        l+=1
                        sheet1.write(r,l,'-')
                        l+=1
                        j="Step1: Power On the ECU\nStep2: Check periodic frame on CANalyzer/CANoe for default values."
                        sheet1.write(r,l,j)
                        l+=1
                        m="1. Capture snapshot of CANalyzer/CANoe window with Valid Data transmitted (default values) to ECU"
                        sheet1.write(r,l,m)
                        r+=1

##To test initial values for each Reception signals
for a,b in txrx.items():
        if b=='Rx':
            for keys, values in CanIdentifier.items():
                if keys==a:
                        naam6.append(a)
                        l=2
                        e+=1
                        testid.append(e)
                        sheet1.write(r,l,e)
                        l+=1
                        n="To test initial values for each Reception signals"
                        sheet1.write(r,l,n)
                        l+=1
                        sheet1.write(r,l,'-')
                        l+=1
                        j="Step1: Power On the ECU\n Step2: Invoke Rte_Read_<p>_<o>(*DataPtr) for each signal  to read initial values.\nStep3: Verify the Data read for each signal with default value.\nwhere <p> is the port name and <o> the VariableDataPrototype within the sender-receiver interface categorizing the port"
                        sheet1.write(r,l,j)
                        l+=1
                        m="1. The read data should match with default configured values."
                        sheet1.write(r,l,m)
                        r+=1

##To test the transmission timeout for the CAN frame with CAN ID
for a,b in txrx.items():
        if b=='Tx':
            for keys, values in CanIdentifier.items():
                if keys==a:
                        naam10.append(a)
                        l=2
                        e+=1
                        testid.append(e)
                        sheet1.write(r,l,e)
                        l+=1
                        n="To test the transmission timeout for the CAN frame with CAN ID ",values,"."
                        sheet1.write(r,l,n)
                        l+=1
                        sheet1.write(r,l,'CAN bus must not be connected.')
                        l+=1
                        j="Step1: Invoke Rte_Write_<p>_<o>(Data) for each signal given below with its value."
                        sheet1.write(r,l,j)
                        l+=1
                        m="1. After timeout value, timeout notification should be triggered."
                        sheet1.write(r,l,m)
                        r+=1

##To test Reception timeout for the CAN frame with CAN ID
for a,b in txrx.items():
        if b=='Rx':
            for keys, values in CanIdentifier.items():
                if keys==a:
                        naam11.append(a)
                        l=2
                        e+=1
                        testid.append(e)
                        sheet1.write(r,l,e)
                        l+=1
                        n="To test Reception timeout for the CAN frame with CAN ID ",values,"."
                        sheet1.write(r,l,n)
                        l+=1
                        sheet1.write(r,l,'CAN bus should be connected.')
                        l+=1
                        j="Step1: Stop sending CAN ID CAN_ID frame from the CANalyzer/CANoe."
                        sheet1.write(r,l,j)
                        l+=1
                        m="1. After timeout value, timeout notification should be triggered."
                        sheet1.write(r,l,m)
                        r+=1


#SHEET2
sheet2=wb.add_sheet('signal')
style = xlwt.easyxf('font: bold 1')
sheet2.write(1, 0, 'TESTCASE NO',style)
sheet2.write(1, 1, 'CAN-ID',style)
sheet2.write(1, 2, 'FRAME-NAME',style)
sheet2.write(1, 3, 'PERIOD_VAL',style)
sheet2.write(1, 4, 'SIGNAL LIST',style)
sheet2.write(1, 5, 'LENGTH',style)
sheet2.write(1, 6, 'MAX',style)
sheet2.write(1, 7, 'MID',style)
sheet2.write(1, 8, 'MIN',style)
sheet2.write(1, 9, 'INT-VALUE',style)
sheet2.write(1, 10, 'TIMEOUT',style)
r=5
e=0
l=0
j=0
nt='N/A'

##To test the Transmission of CAN frame for CAN ID,with Signal Min/Mid/Max values and with Periodicity
for i in naam:
    l=0
    j+=1
    for a,b in CanIdentifier.items():
        if i==a:
            sheet2.write(r,l,j)
            l+=1
            sheet2.write(r,l,b)
            l+=1
            sheet2.write(r,l,i)
            l+=1
    for c,d in timeper.items():
        if i==c:
            sheet2.write(r,l,d)
            l+=1
            for e,f in cane.items():
                if e==i:
                    for signal in f:
                        n=l
                        sheet2.write(r,n,signal)
                        n+=1
                        for g,h in lengsignal.items():
                            if signal==g:
                                p=int(h)
                                sheet2.write(r,n,p)
                                n+=1
                                if p==1:
                                    k=1
                                else:
                                    k=(2**p)
                                sheet2.write(r,n,k)
                                n+=1

                                if p==1:
                                    mid=1
                                else:
                                    mid=k/2
                                sheet2.write(r,n,mid)
                                n+=1
                                mini=0
                                sheet2.write(r,n,mini)
                                n+=1
                                sheet2.write(r,n,nt)
                        for ca,ta in timeoutee.items():
                            if signal==ca:
                                n+=1
                                sheet2.write(r,n,nt)
                                r+=1

##To test the Reception of CAN frame for CAN ID,with signal Min/Mid/Max values and with Periodicity
for i in naam2:
    l=0
    j+=1
    for a,b in CanIdentifier.items():
        if i==a:
            sheet2.write(r,l,j)
            l+=1
            sheet2.write(r,l,b)
            l+=1
            sheet2.write(r,l,i)
            l+=1
            d='N/A'
            sheet2.write(r,l,d)
            l+=1
            for e,f in cane.items():
                if e==i:
                    for signal in f:
                        n=l
                        sheet2.write(r,n,signal)
                        n+=1
                        for g,h in lengsignal.items():
                            if signal==g:
                                if h!='N/A':
                                    p=int(h)
                                    sheet2.write(r,n,p)
                                    n+=1
                                    if p==1:
                                        k=1
                                    else:
                                        k=(2**p)
                                    sheet2.write(r,n,k)
                                    n+=1
                                    if p==1:
                                        mid=1
                                    else:
                                        mid=k/2
                                    sheet2.write(r,n,mid)
                                    n+=1
                                    mini=0
                                    sheet2.write(r,n,mini)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                else:
                                    p=h
                                    sheet2.write(r,n,p)
                                    n+=1
                                    sheet2.write(r,n,h)
                                    n+=1
                                    sheet2.write(r,n,h)
                                    n+=1
                                    sheet2.write(r,n,h)
                                    n+=1
                                    sheet2.write(r,n,nt)
                        for ca,ta in timeoutee.items():
                            if signal==ca:
                                n+=1
                                sheet2.write(r,n,nt)
                                r+=1

##To test the Transmission of CAN frame for CAN ID,with signal value when user requested from CANalyzer/CANoe.
for i in naam3:
    l=0
    j+=1
    for a,b in CanIdentifier.items():
        if i==a:
            sheet2.write(r,l,j)
            l+=1
            sheet2.write(r,l,b)
            l+=1
            sheet2.write(r,l,i)
            l+=1

            d='N/A'
            sheet2.write(r,l,d)
            l+=1
            for e,f in cane.items():
                if e==i:
                    for signal in f:
                        n=l
                        sheet2.write(r,n,signal)
                        n+=1
                        for g,h in lengsignal.items():
                            if signal==g:
                                if h!='N/A':
                                    p=int(h)
                                    sheet2.write(r,n,p)
                                    n+=1
                                    k=(2**p)
                                    sheet2.write(r,n,k)
                                    n+=1
                                    mid=k/2
                                    sheet2.write(r,n,mid)
                                    n+=1
                                    mini=0
                                    sheet2.write(r,n,mini)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                else:
                                    p=h
                                    sheet2.write(r,n,p)
                                    n+=1
                                    sheet2.write(r,n,h)
                                    n+=1

                                    sheet2.write(r,n,h)
                                    n+=1
                                    sheet2.write(r,n,h)
                                    n+=1
                                    sheet2.write(r,n,nt)
                        for ca,ta in timeoutee.items():
                            if signal==ca:
                                n+=1
                                sheet2.write(r,n,ta)
                                r+=1

##To test initial values for each Transmission signals
for i in naam4:
    l=0
    j+=1
    for a,b in CanIdentifier.items():
        if i==a:
            sheet2.write(r,l,j)
            l+=1
            sheet2.write(r,l,b)
            l+=1
            sheet2.write(r,l,i)
            l+=1
            d='N/A'
            sheet2.write(r,l,d)
            l+=1
            for e,f in cane.items():
                if e==i:
                    for signal in f:
                        n=l
                        sheet2.write(r,n,signal)
                        n+=1
                        for g,h in initvalue.items():
                            if signal==g:
                                if h!='N/A':
                                    p=int(h)
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    mini=0
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,p)
                                else:
                                    p=h
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,p)
                        for ca,ta in timeoutee.items():
                            if signal==ca:
                                n+=1
                                sheet2.write(r,n,ta)
                                r+=1

##To test initial values for each Reception signals
for i in naam6:
    l=0
    j+=1
    for a,b in CanIdentifier.items():
        if i==a:
            sheet2.write(r,l,j)
            l+=1
            sheet2.write(r,l,b)
            l+=1
            sheet2.write(r,l,i)
            l+=1

            d='N/A'
            sheet2.write(r,l,d)
            l+=1
            for e,f in cane.items():
                if e==i:
                    for signal in f:
                        n=l
                        sheet2.write(r,n,signal)
                        n+=1
                        for g,h in initvalue.items():
                            if signal==g:
                                if h!='N/A':
                                    p=int(h)
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    mini=0
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,p)
                                else:
                                    p=h
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,p)
                        for ca,ta in timeoutee.items():
                            if signal==ca:
                                n+=1
                                sheet2.write(r,n,ta)
                                r+=1

##To test the transmission timeout for the CAN frame with CAN ID
c=0
for i in naam10:
    l=0
    j+=1
    if c==1:
        r+=1
    for a,b in CanIdentifier.items():
        if i==a:
            sheet2.write(r,l,j)
            l+=1
            sheet2.write(r,l,b)
            l+=1
            sheet2.write(r,l,i)
            l+=1
            d='N/A'
            sheet2.write(r,l,d)
            l+=1
            for e,f in cane.items():
                if e==i:
                    for signal in f:
                        c=1
                        n=l
                        for g,h in timeoutee.items():
                            if signal==g:
                                if h!='N/A':
                                    sheet2.write(r,n,signal)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    mini=0
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,h)
                                    r+=1
                                    c=0


##To test Reception timeout for the CAN frame with CAN ID
if c==1:
    r+=1
c=0
for i in naam11:
    l=0
    j+=1
    if c==1:
        r+=1
    for a,b in CanIdentifier.items():
        if i==a:
            sheet2.write(r,l,j)
            l+=1
            sheet2.write(r,l,b)
            l+=1
            sheet2.write(r,l,i)
            l+=1
            d='N/A'
            sheet2.write(r,l,d)
            l+=1
            for e,f in cane.items():
                if e==i:
                    for signal in f:
                        n=l
                        c=1
                        for g,h in timeoutee.items():
                            if signal==g:
                                if h!='N/A':
                                    sheet2.write(r,n,signal)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    mini=0
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,nt)
                                    n+=1
                                    sheet2.write(r,n,h)
                                    r+=1
                                    c=0




#savedto=s=r"ITP_HFS.xls"
wb.save(OUTITP)
print('ITP Output saved to Excel')

#TEST APP
file = open(OUTHEAD,'w')
with file as f:
   for keys, values in cane.items():
       cann='typedef '+'struct ' + keys
       f.write(cann)
       f.write('\n')
       f.write('{')
       f.write('\n')
#        print(cann,'\n{')

       for a in values:
           for c,d in leng.items():
               if a==c:
                   d=int(d)
                   if d<=8:
                       uni='uint8 ' + c+';'
                       f.write(uni)
                       f.write('\n')
#                        print('unit8',c)
                   if d<=16 and d>=9:
                       uni='uint16 ' + c+';'
                       f.write(uni)
                       f.write('\n')
#                        print('unit16',c)
                   if d<=32 and d>=17:
                       uni='uint32 ' + c +';'
                       f.write(uni)
                       f.write('\n')
#                        print('unit32',c)
                   if d<=64 and d>=33:
                       uni='uint64 ' + c +';'
                       f.write(uni)
                       f.write('\n')
#                        print('unit64',c)
       ple='}Incedence_'+keys+';'
       f.write(ple)
       f.write('\n')
       f.write('\n')
#        print('}\n\n')

file1 = open(OUTC,'w')
with file1 as f:
   for keys, values in cane.items():
       clae='Incedence_'+keys+ ' Test_'+keys +'[3]={{'
       f.write(clae)
       minie=[]
       maxie=[]
       midie=[]
       for a in values:
           for c,d in leng.items():
               if a==c:
                   d=int(d)
                   m=0
                   minie.append(m)
                   i=(2**d)/2
                   midie.append(int(i))
                   x=((2**d)-1)
                   maxie.append(int(x))
       two=0
       tre=0
       four=0
       for da in minie:
           one=len(minie)
           f.write(str(da))
           two+=1
           if one!=two:
               f.write(',')
       f.write('},')
       f.write('\n')
       f.write('{')
       for da in midie:
           one=len(midie)
           da=str(da)
           f.write(da)
           tre+=1
           if one!=tre:
               f.write(',')
       f.write('},')
       f.write('\n')
       f.write('{')
       for da in maxie:
           one=len(midie)
           da=str(da)
           f.write(da)
           four+=1
           if one!=four:
               f.write(',')
       f.write('}};')
       f.write('\n')
       f.write('\n')
print('header and c file generated')
#messagebox.showinfo( "KPIT", "Output generated")

#OUTMAIN
file2 = open(OUTMAIN,'w')

case=0
txfra=[]
for k,v in txrx.items():
    if v=='Tx':
        txfra.append(k)
with file2 as f:
    f.write('#include \"Test_App.h\"\n#if(DCU_TESTAPP == ON)\n#include \"Test_App_Data.h\"\n#include \"Com_Cfg.h\"\n#include \"Rte_Type.h\" ')
    f.write('\n\n\n')
    f.write('void Test_Case_Tx_FS(void)')
    f.write('\n')
    f.write('{')
    f.write('\n\n')
    f.write('\t')
    f.write('uint8 Gaa_Com_TestResult[Total_ComTest_Count][3];')
    f.write('\n')
    f.write('\t')
    f.write('uint8 Lucrange = 0;')
    f.write('\n')
    f.write('\t')
    f.write('static uint8 LucTestCase_No = 1;')
    f.write('\n')
    f.write('\t')
    f.write('uint8 LaaTc_1[5];')
    f.write('\n\n')
    f.write('\t\t')
    f.write('switch(LucTestCase_No)')
    f.write('\n')
    f.write('\t\t')
    f.write('{')
    f.write('\n')


    for keys, values in cane.items():
        for txe in txfra:
            if txe==keys:
                case+=1
                cann='case '+ str(case) +':'
                f.write('\t\t\t')
                f.write(cann)
                f.write('\n\n')
                vt=-1
                for a in values:
                    vt+=1
                    lact='LaaTc_1['+str(vt)+']= Com_SendSignal(&Test_'+keys+'[Lucrange].'+a+');'
                    f.write('\t\t\t\t')
                    f.write(lact)
                    f.write('\n')
  ##               ple='}Incedence_'+keys+';'
  ##               f.write(ple)
                f.write('\n\n')
                nonv=-1
                motte=''
                for i in range(0,vt):
                    if i!=vt:
                        lotte='(LaaTc_1['+str(i)+'] == E_OK) && '
                        motte= motte +lotte
                lotte='(LaaTc_1['+str(vt)+'] == E_OK)'
                motte= motte +lotte
                cole= 'if('+motte+')'
                f.write('\t\t\t\t\t')
                f.write(cole)
                f.write('\n')
                f.write('\t\t\t\t\t\t')
                f.write('{')
                f.write('\n')
                f.write('\t\t\t\t\t\t\t')
                clit='Gaa_Com_TestResult[LucTestCase_No][Lucrange] = PASS;'
                f.write(clit)
                f.write('\n')
                f.write('\t\t\t\t\t\t')
                f.write('}')
                f.write('\n')
                f.write('\t\t\t\t\t')
                f.write('else')
                f.write('\n')
                f.write('\t\t\t\t\t\t')
                f.write('{')
                f.write('\n')
                f.write('\t\t\t\t\t\t\t')
                flit='Gaa_Com_TestResult[LucTestCase_No][Lucrange] = FAIL;'
                f.write(flit)
                f.write('\n')
                f.write('\t\t\t\t\t\t')
                f.write('}')
                f.write('\n')
                f.write('\t\t\t\t\t\t')
                f.write('break')
                f.write('\n\n')

                f.write('\n')
    f.write('\t\t')
    f.write('}')
    f.write('\n')
    f.write('}')
    f.write('#endif')
#MIN CAPL
file3 = open(CAPLMIN,'w')
with file3 as f:
  for keys, values in cane.items():
        for txe in txfra:
            if txe==keys:
                case+=1
                caple='on message '+ keys
                f.write('\n')
                f.write(caple)
                f.write('\n')
                f.write('{')
                f.write('\n')
                motte=''
                for i in range(len(values)):
                  if i!=len(values):
                    pep=values[i]
                    lotte='(getsignal('+pep+') == 0) && '
                    motte= motte +lotte
                lotte='(getsignal('+ pep +') == 0)'
                motte= motte +lotte
                cole= 'if('+motte+')'
                f.write('\t')
                f.write(cole)
                f.write('\n')
                f.write('\t\t')
                f.write('{')
                f.write('\n')
                f.write('\t\t')
                f.write('write(\"PASS\")')
                f.write('\n')
                f.write('\t\t')
                f.write('}')
                f.write('\n')
                f.write('\t')
                f.write('else')
                f.write('\n')
                f.write('\t\t')
                f.write('{')
                f.write('\n')
                f.write('\t\t')
                f.write('write(\"FAIL\")')
                f.write('\n')
                f.write('\t\t')
                f.write('}')
                f.write('\n')
                f.write('}')
                f.write('\n\n')

#MAX CAPL
file4 = open(CAPLMAX,'w')
with file4 as f:
  for keys, values in cane.items():
        for txe in txfra:
            if txe==keys:
                case+=1
                caple='on message '+ keys
                f.write('\n')
                f.write(caple)
                f.write('\n')
                f.write('{')
                f.write('\n')
                motte=''
                for i in range(len(values)):
                  pep=values[i]
                  for c,d in leng.items():
                    if pep==c:
                      x=hex((2**int(d))-1)
                      if i!=len(values):
                        lotte='(getsignal('+pep+') =='+ str(x)+') && '
                        motte= motte +lotte
                      elif i==len(values):
                        lotte='(getsignal('+ pep +') == '+str(x)+')'
                        motte= motte +lotte
                cole= 'if('+motte+')'
                f.write('\t')
                f.write(cole)
                f.write('\n')
                f.write('\t\t')
                f.write('{')
                f.write('\n')
                f.write('\t\t')
                f.write('write(\"PASS\")')
                f.write('\n')
                f.write('\t\t')
                f.write('}')
                f.write('\n')
                f.write('\t')
                f.write('else')
                f.write('\n')
                f.write('\t\t')
                f.write('{')
                f.write('\n')
                f.write('\t\t')
                f.write('write(\"FAIL\")')
                f.write('\n')
                f.write('\t\t')
                f.write('}')
                f.write('\n')
                f.write('}')
                f.write('\n\n')

#MID CAPL
file4 = open(CAPLMID,'w')
with file4 as f:
  for keys, values in cane.items():
        for txe in txfra:
            if txe==keys:
                case+=1
                caple='on message '+ keys
                f.write('\n')
                f.write(caple)
                f.write('\n')
                f.write('{')
                f.write('\n')
                motte=''
                for i in range(len(values)):
                  pep=values[i]
                  for c,d in leng.items():
                    if pep==c:
                      x=int((2**int(d))/2)
                      if i!=len(values):
                        lotte='(getsignal('+pep+') =='+ str(x)+') && '
                        motte= motte +lotte
                      elif i==len(values):
                        lotte='(getsignal('+ pep +') == '+str(x)+')'
                        motte= motte +lotte
                cole= 'if('+motte+')'
                f.write('\t')
                f.write(cole)
                f.write('\n')
                f.write('\t\t')
                f.write('{')
                f.write('\n')
                f.write('\t\t')
                f.write('write(\"PASS\")')
                f.write('\n')
                f.write('\t\t')
                f.write('}')
                f.write('\n')
                f.write('\t')
                f.write('else')
                f.write('\n')
                f.write('\t\t')
                f.write('{')
                f.write('\n')
                f.write('\t\t')
                f.write('write(\"FAIL\")')
                f.write('\n')
                f.write('\t\t')
                f.write('}')
                f.write('\n')
                f.write('}')
                f.write('\n\n')


#RX ONLY ##########################################################################################################
#TEST APP
rxfra=[]
for k,v in txrx.items():
    if v=='Rx':
        rxfra.append(k)
file5 = open(RX1,'w')
with file5 as f:
   for keys, values in cane.items():
       for rxe in rxfra:
            if rxe==keys:
               cann='typedef '+'struct ' + keys
               f.write(cann)
               f.write('\n')
               f.write('{')
               f.write('\n')
        #        print(cann,'\n{')

               for a in values:
                   for c,d in leng.items():
                       if a==c:
                           d=int(d)
                           if d<=8:
                               uni='uint8 ' + c+';'
                               f.write(uni)
                               f.write('\n')
        #                        print('unit8',c)
                           if d<=16 and d>=9:
                               uni='uint16 ' + c+';'
                               f.write(uni)
                               f.write('\n')
        #                        print('unit16',c)
                           if d<=32 and d>=17:
                               uni='uint32 ' + c +';'
                               f.write(uni)
                               f.write('\n')
        #                        print('unit32',c)
                           if d<=64 and d>=33:
                               uni='uint64 ' + c +';'
                               f.write(uni)
                               f.write('\n')
        #                        print('unit64',c)
               ple='}Incedence_'+keys+';'
               f.write(ple)
               f.write('\n')
               f.write('\n')
        #        print('}\n\n')

file6 = open(RX2,'w')
with file6 as f:
          for keys, values in cane.items():
              for rxe in rxfra:
                  if rxe==keys:
                       clae='Incedence_'+keys+ '\n Test_'+keys +'[3]={{'
                       f.write(clae)
                       minie=[]
                       maxie=[]
                       midie=[]
                       for a in values:
                           for c,d in leng.items():
                               if a==c:
                                   d=int(d)
                                   m=0
                                   minie.append(m)
                                   i=(2**d)/2
                                   midie.append(int(i))
                                   x=((2**d)-1)
                                   maxie.append(int(x))
                       two=0
                       tre=0
                       four=0
                       for da in minie:
                           one=len(minie)
                           f.write(str(da))
                           two+=1
                           if one!=two:
                               f.write(',')
                       f.write('},')
                       f.write('\n')
                       f.write('{')
                       for da in midie:
                           one=len(midie)
                           da=str(da)
                           f.write(da)
                           tre+=1
                           if one!=tre:
                               f.write(',')
                       f.write('},')
                       f.write('\n')
                       f.write('{')
                       for da in maxie:
                           one=len(midie)
                           da=str(da)
                           f.write(da)
                           four+=1
                           if one!=four:
                               f.write(',')
                       f.write('}};')
                       f.write('\n')
                       f.write('\n')
print('header and c file generated')
#messagebox.showinfo( "KPIT", "Output generated")

#OUTMAIN
file7 = open(RX3,'w')

case=0
with file7 as f:
    f.write('#include \"Test_App.h\"\n#if(DCU_TESTAPP == ON)\n#include \"Test_App_Data.h\"\n#include \"Com_Cfg.h\"\n#include \"Rte_Type.h\" ')
    f.write('\n\n\n')
    f.write('void Test_Case_Tx_FS(void)')
    f.write('\n')
    f.write('{')
    f.write('\n\n')
    f.write('\t')
    f.write('uint8 Gaa_Com_TestResult[Total_ComTest_Count][3];')
    f.write('\n')
    f.write('\t')
    f.write('uint8 Lucrange = 0;')
    f.write('\n')
    f.write('\t')
    f.write('static uint8 LucTestCase_No = 1;')
    f.write('\n')
    f.write('\t')
    f.write('uint8 LaaTc_1[5];')
    f.write('\n\n')
    f.write('\t\t')
    f.write('switch(LucTestCase_No)')
    f.write('\n')
    f.write('\t\t')
    f.write('{')
    f.write('\n')
    for keys, values in cane.items():
        for rxe in rxfra:
            if rxe==keys:
                laca=len(values)
                colaaa=''
                picaaa=laca
                picaaa-=1
                i=0
                case+=1
                f.write('\t\t\t')
                f.write('uint8 LaaTc_'+str(case)+'['+str(picaaa)+']')
                f.write('\n')
                for  i in range(0,laca):
                    if i<=laca:
                        f.write('\t\t\t')
                        loso='uint8 LucRx_Tc'+str(case)+'_Data_'+str(i)+';'
                        f.write(loso)
                        f.write('\n')
                f.write('\n')
                f.write('\n')



    case=0
    for keys, values in cane.items():
        for rxe in rxfra:
            if rxe==keys:
                laca=len(values)
                colaaa=''
                picaaa=laca
                picaaa-=1
                i=0
                case+=1
                for  i in range(0,laca):
                    if i<=laca:
                        loso='uint8 LucRx_Tc'+str(case)+'_Data_'+str(i)+';'

                cann='case '+ str(case) +':'
                f.write('\t\t\t')
                f.write(cann)
                f.write('\n\n')
                vt=-1
                for a in values:
                    vt+=1
                    lact='LaaTc_'+str(case)+'['+str(vt)+']= Rte_Read_Dummy_SignalProcessor_Rx_PpPdu_'+a+'(LucRx_Tc'+str(case)+'_Data_'+str(vt)+');'
                    f.write('\t\t\t\t')
                    f.write(lact)
                    f.write('\n')
                    if vt!=picaaa:
                        molaaa='(LucRx_Tc'+str(case)+'_Data_'+str(vt)+')== Test_'+keys+'[Lucrange].'+a+' && '
                    if vt==picaaa:
                        molaaa='(LucRx_Tc'+str(case)+'_Data_'+str(vt)+')== Test_'+keys+'[Lucrange].'+a
                    colaaa=colaaa+molaaa
  ##               ple='}Incedence_'+keys+';'
  ##               f.write(ple)
                f.write('\n\n')
                nonv=-1
                motte=''
                for i in range(0,vt):
                    if i!=vt:
                        lotte='(LaaTc_1['+str(i)+'] == E_OK) && '
                        motte= motte +lotte
                lotte='(LaaTc_1['+str(vt)+'] == E_OK)'
                motte= motte +lotte
                cole= 'if('+motte+')'
                f.write('\t\t\t\t\t')
                f.write(cole)
                f.write('\n')
                f.write('\t\t\t\t\t\t')
                f.write('{')
                f.write('\n')
                f.write('\t\t\t\t\t\t\t')
                colaa='if('+colaaa+')'
                f.write(colaa)
                f.write('\n')
                f.write('\t\t\t\t\t\t\t')
                f.write('{')
                f.write('\n')
                f.write('\t\t\t\t\t\t\t\t')
                clit='Gaa_Com_TestResult[LucTestCase_No][Lucrange] = PASS;'
                f.write(clit)
                f.write('\n')
                f.write('\t\t\t\t\t\t\t')
                f.write('}')
                f.write('\n')
                f.write('\t\t\t\t\t\t\t')
                f.write('else')
                f.write('\n')
                f.write('\t\t\t\t\t\t\t')
                f.write('{')
                f.write('\n')
                f.write('\t\t\t\t\t\t\t\t')
                plit='Gaa_Com_TestResult[LucTestCase_No][Lucrange] = FAIL;'
                f.write(plit)
                f.write('\n')
                f.write('\t\t\t\t\t\t\t')
                f.write('}')
                f.write('\n')
                f.write('\t\t\t\t\t\t')
                f.write('}')
                f.write('\n')
                f.write('\t\t\t\t\t')
                f.write('else')
                f.write('\n')
                f.write('\t\t\t\t\t\t')
                f.write('{')
                f.write('\n')
                f.write('\t\t\t\t\t\t\t')
                flit='Gaa_Com_TestResult[LucTestCase_No][Lucrange] = FAIL;'
                f.write(flit)
                f.write('\n')
                f.write('\t\t\t\t\t\t')
                f.write('}')
                f.write('\n')
                f.write('\t\t\t\t\t\t')
                f.write('break')
                f.write('\n\n')

                f.write('\n')
    f.write('\t\t')
    f.write('}')
    f.write('\n')
    f.write('}')
    f.write('#endif')




file8 = open(RX4,'w')

with file8 as f:
    f.write('#if !defined(DATA_H)\n# define DATA_H" ')
    f.write('\n\n')
    f.write('#if(DCU_TESTAPP == ON)\n#include \"Std_Types.h\"\n#include \"Rte_Type.h\"\n#include \"Test_App.h\"')
    f.write('\n\n')
    for keys, values in cane.items():
        for rxe in rxfra:
            if rxe==keys:
                pktr='extern Incedence_'+keys+' Test_'+keys+'[3];'
                f.write(pktr)
                f.write('\n')
    f.write('\n\n')
    f.write('#endif')



##
##
##
##
##
##
##
##
##
##
##
